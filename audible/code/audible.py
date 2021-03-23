
#  Python imports
import threading
import xlsxwriter 
import json, time
import pandas as pd
from numpy import nan
import openpyxl as op
from bs4 import BeautifulSoup
from selenium import webdriver 


book_prefix = {
    'Australia':'https://www.amazon.com.au',
    'Brazil':'https://www.amazon.com.br',
    'Canada':'https://www.amazon.ca',
    'China':'https://www.amazon.cn',
    'France':'https://www.amazon.fr',
    'Germany':'https://www.amazon.de',
    'India':'https://www.amazon.in',
    'Italy':'https://www.amazon.it',
    'Japan':'https://www.amazon.co.jp',
    'Mexico':'https://www.amazon.com.mx',
    'Netherlands':'https://www.amazon.nl',
    'Ploand':'https://www.amazon.pl',
    # 'Saudi Arabia':'https://www.amazon.',
    'Singapore':'https://www.amazon.sg',
    'Spain':'https://www.amazon.es',
    'Sweden':'https://www.amazon.se',
    'Turkey':'https://www.amazon.tr',
    'United Arab Emirates':'https://www.amazon.ae',
    'United States' : 'https://www.amazon.com',
    'United Kingdom': 'https://www.amazon.co.uk'
}


class audible:
    count = 1
    country = ''
    sub_level = 0
    sub_names = {}
    categories = []
    book_number = 50
    audible_filename = ''
    audible_categories = {}
    data_fields =  ['category', 'subcat-1', 'subcat-2', 'subcat-3', 'subcat-4']

    def __init__(self, country):
        self.country = country
        self.audible_filename = f'../data/audible_{country}.xlsx'
        setting =  pd.read_excel('settings.xlsx', 'settings')
        datas = setting['audible-data-fields']
        cats  = setting['audible-categories']
        self.book_number = int(setting['book-number'][0])
        for i in datas.index:
            if datas[i] is not nan:    self.data_fields.append(datas[i])
        for i in cats.index:
            if cats[i] is not nan:     self.categories.append(cats[i])
        self.audible_categories = self.update_category_list(f'../category_list/{country}.json')
        print(country)
        print(self.audible_categories)

    def update_category_list(self, filename):
        with open(filename, 'r+') as read_file:
            return json.load(read_file)
    
    def scrape_category(self):
        count = 1
        for cat in self.categories:
            self.create_excel_file(cat, self.audible_filename)
            break
        for category in self.categories:   # Create new thread for category
            self.browser = webdriver.Chrome('../../chromedriver.exe') 
            self.sub_level = 1
            self.sub_names = {'category': category, 'subcat-1' : 'null', 'subcat-2' : 'null', 'subcat-3' : 'null', 'subcat-4' : 'null'}
            workbook = op.load_workbook(self.audible_filename, False)
            try:
                workbook[category]
            except:
                workbook.create_sheet(category)
                worksheet = workbook[category]
                worksheet.append(self.headers())
                workbook.save(self.audible_filename)
                workbook.close()
            try:
                subcategories = self.audible_categories[ category ]
                t = threading.Thread(target=self.helper_category_books, args=(subcategories, self.audible_filename, ))
                t.start()
                t.join()
            except:
                print ("Error: unable to start new thread")
            count += 1
            self.browser.quit()


    def helper_category_books(self, subcategories, filename):
        for subcat in subcategories:
            if not type(subcategories[subcat]) is dict:
                self.update_subnames(subcat)
                try:    # Create new thread for category
                    t = threading.Thread(target=self.category_books, args=(filename, subcategories[subcat], ))
                    t.start()
                    t.join()
                except:
                    print ("Error: unable to start new thread")
            else:
                self.update_subnames(subcat)
                self.sub_level += 1
                self.helper_category_books(subcategories[subcat], filename)
        self.update_subnames('null')
        self.sub_level -= 1

    def update_subnames(self, subcat):
        print('\n' ,self.sub_level, ' - ', self.sub_names)
        if self.sub_level == 1: self.sub_names['subcat-1'] = subcat
        if self.sub_level == 2: self.sub_names['subcat-2'] = subcat
        if self.sub_level == 3: self.sub_names['subcat-3'] = subcat
        if self.sub_level == 4: self.sub_names['subcat-4'] = subcat   
        
    def category_books(self, filename, link):
        books = []  # for book-data
        self.browser.set_window_position(500,0)
        try:
            self.browser.get(link)
        except:
            print('Failed to Load')
            return False
        source = self.browser.page_source
        soup = BeautifulSoup(source, features='lxml')
        book_sections = soup.find_all('div', attrs={'class':'a-section a-spacing-none aok-relative'})
        book_count = 1
        for book in book_sections:
            book_count += 1
            a_tags = book.find_all('a', attrs={'class':'a-link-normal'})
            book_details_link = book_prefix[self.country] + a_tags[0]['href']
            try:
                self.browser.get(book_details_link)
                source = self.browser.page_source
                soup = BeautifulSoup(source, features='lxml')
                
                title = soup.find('span', attrs={'id':'productTitle'}).get_text().strip('\n')
                span = soup.find_all('span', attrs={'class', 'author notFaded'})
                author = span[0].find('a', attrs={'class':'a-link-normal'}).get_text()
                rating = soup.find('span', attrs={'id':'acrCustomerReviewText'}).get_text().replace(' ratings', '')
                stars = soup.find('span', attrs={'class': 'reviewCountTextLinkedHistogram noUnderline'})['title'].replace(' out of 5 stars', '')
                table = soup.find('table', attrs={'class':'a-keyvalue a-vertical-stripes a-span6'})
                table = table.find('tbody')
                tr_tags = table.find_all('tr')
                tr_list = []
                for i in range(len(tr_tags)-1):
                    tr_list.append(tr_tags[i])

                details = {}
                for name in self.sub_names: details[name] = self.sub_names[name]
                if 'Title'          in self.data_fields:    details['Title']        = title
                if 'Web-Link'       in self.data_fields:    details['Web-Link']     = book_details_link
                if 'Author'         in self.data_fields:    details['Author']       = author
                if 'Ratings'        in self.data_fields:    details['Ratings']      = rating
                if 'Stars'          in self.data_fields:    details['Stars']        = stars

                for tr in tr_list:
                    span = tr.find('th')
                    span = span.find('span')
                    heading = span.get_text() 
                    data = tr.find('td')
                    try:    data = data.find('span').get_text()
                    except: data = data.find('a').get_text()
                    details[heading] = data

                bst_heading = table.find('th', attrs={'class':'a-color-secondary a-size-base prodDetSectionEntry'}).get_text().replace('\n','')
                bst_tr = None
                for tr in tr_tags:  bst_tr = tr
                td = bst_tr.find('td')
                span = td.find('span')
                spans = span.find_all('span')
                bst_data = []
                for span in spans:
                    try:    bst_data.append(span.get_text().split('(')[0])
                    except:    bst_data.append(span.get_text())
                details[bst_heading] = bst_data
                # print( 'A-', self.count, '. ', details, '\n')
                print(self.count, end=" ")
                self.count += 1
                books.append(details)
            except: 
                continue
            if book_count > self.book_number:   break
        # Saving in Excel File  
        self.write_to_excel(self.audible_filename, books)

    def create_excel_file(self, category_name, filename):
        # creating new excle file
        workbook = xlsxwriter.Workbook(filename)
        workbook.add_worksheet(category_name)
        workbook.close()
        workbook = op.load_workbook(filename, False)
        worksheet = workbook[category_name]
        worksheet.append(self.headers())
        workbook.save('../data/' + filename)
        workbook.close()

    def write_to_excel(self, filename, books=[]):
        workbook = op.load_workbook(filename, False)
        worksheet = workbook[self.sub_names['category']]
        for book in books:
            data = []
            for data_field in self.data_fields:
                try:
                    if data_field == 'Best Sellers Rank':
                        for v in book[data_field]: data.append(v)
                    else:   data.append(book[data_field])
                except: data.append('N/A')
            worksheet.append(data)
        workbook.save(filename)
        workbook.close()

    def headers(self):
        header = []
        for data_field in self.data_fields:
            header.append(data_field)
        return header


def selected_countries():
    countries =  pd.read_excel('../category_list/countries.xlsx', 'countries')
    datas = countries['countries']
    country_list = []
    for i in datas.index:
            if datas[i] is not nan:    
                country_list.append(datas[i])
    return country_list

if __name__ == '__main__':
    countries = selected_countries()
    print(countries)
    for country in countries:
        try:
            audi = audible(country)
            audi.scrape_category()
            time.sleep(5)
        except:
            print(country, '-List Not Found')
