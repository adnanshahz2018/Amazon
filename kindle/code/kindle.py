
#  Python imports
import threading
import xlsxwriter 
import json, time
import os, random
import pandas as pd
from numpy import nan
import openpyxl as op
from bs4 import BeautifulSoup
from selenium import webdriver 


book_prefix = {
    'Australia':'https://www.amazon.com.au',
    'Brazil':'https://www.amazon.com.br',
    'Canada':'https://www.amazon.ca',
    'China':'https://www.amazon.cn',
    'France':'https://www.amazon.fr',
    'Germany':'https://www.amazon.de',
    'India':'https://www.amazon.in',
    'Italy':'https://www.amazon.it',
    'Japan':'https://www.amazon.co.jp',
    'Mexico':'https://www.amazon.com.mx',
    'Netherlands':'https://www.amazon.nl',
    'Ploand':'https://www.amazon.pl',
    # 'Saudi Arabia':'https://www.amazon.',
    'Singapore':'https://www.amazon.sg',
    'Spain':'https://www.amazon.es',
    'Sweden':'https://www.amazon.se',
    'Turkey':'https://www.amazon.tr',
    'United Arab Emirates':'https://www.amazon.ae',
    'United States' : 'https://www.amazon.com',
    'United Kingdom': 'https://www.amazon.co.uk'
}

class kindle:
    count = 0
    country = ''
    sub_level = 0
    sub_names = {}
    categories = []
    book_number = 50
    work_done = list()
    kindle_categories = {}
    kindle_filename = ''
    data_fields =  ['category', 'subcat-1', 'subcat-2', 'subcat-3', 'subcat-4']
    
    def __init__(self, country='United States'):
        self.country = country
        self.kindle_filename = f'../data/kindle_{country}.xlsx'
        setting =  pd.read_excel('settings.xlsx', 'settings')
        datas = setting['kindle-data-fields']
        cats  = setting['kindle-categories']
        self.book_number = int(setting['book-number'][0])
        for i in datas.index:
            if datas[i] is not nan:    self.data_fields.append(datas[i])
        for i in cats.index:
            if cats[i] is not nan:     self.categories.append(cats[i])
        self.kindle_categories = self.update_category_list(f'../category_list/{country}.json', country)
        if self.kindle_categories:
            print(country)
            print(self.kindle_categories)

    def update_category_list(self, filename, country):
        if not os.path.exists(filename):
            print(f'{country}.json File NOT Foundt')
            return False
        else:
            with open(filename, 'r+') as read_file:
                return json.load(read_file)

    def scrape_category(self):
        for cat in self.categories:
            self.create_excel_file(cat, self.kindle_filename)
            break
        workbook = op.load_workbook(self.kindle_filename, False)
        for category in self.categories:   # Create new thread for category
            if category in self.kindle_categories:
                browser = webdriver.Chrome('../../chromedriver.exe') 
                self.work_done.append(False)
                self.sub_level = 1
                self.sub_names = {'category': category, 'subcat-1' : 'null', 'subcat-2' : 'null', 'subcat-3' : 'null', 'subcat-4' : 'null'}
                try:
                    workbook[category]
                except:
                    workbook.create_sheet(category)
                    worksheet = workbook[category]
                    worksheet.append(self.headers())
                    workbook.save(self.kindle_filename)
                    workbook.close()
                try:
                    subcategories = self.kindle_categories[ category ]
                    t = threading.Thread(target=self.intermediate, args=(self.count, browser, subcategories, ))
                    t.start()
                    time.sleep(random.randint(5, 10))
                    # t.join()
                except:
                    print ("Error: unable to start new thread")
                self.count += 1
                # print(self.country, '- count  == ', count)
    
    def update_subnames(self, subcat):
        print('\n' ,self.sub_level, ' - ', self.sub_names)
        if self.sub_level == 1: self.sub_names['subcat-1'] = subcat
        if self.sub_level == 2: self.sub_names['subcat-2'] = subcat
        if self.sub_level == 3: self.sub_names['subcat-3'] = subcat
        if self.sub_level == 4: self.sub_names['subcat-4'] = subcat   

    def intermediate(self, count, browser, subcategories):
        self.helper_category_books(browser, subcategories)
        # print('\n\n Category Finished -------\n\n')
        self.work_done[count] = True
        browser.quit()

    def helper_category_books(self, browser, subcategories):
        for subcat in subcategories:
            if not type(subcategories[subcat]) is dict:
                self.update_subnames(subcat)
                try:    # Create new thread for category
                    t = threading.Thread(target=self.category_books, args=(browser, subcategories[subcat], ))
                    t.start()
                    t.join()
                except:
                    print ("Error: unable to start new thread")
            else:
                self.update_subnames(subcat)
                self.sub_level += 1
                self.helper_category_books(browser, subcategories[subcat])
        self.update_subnames('null')
        self.sub_level -= 1
    
    def category_books(self, browser, link):
        books = []  # for book-data
        browser.set_window_position(500,0)
        try:
            browser.get(link)
        except:
            print('Failed to Load')
            return False
        source = browser.page_source
        soup = BeautifulSoup(source, features='lxml')
        book_sections = soup.find_all('div', attrs={'class':'a-section a-spacing-none aok-relative'})
        book_count = 1
        for book in book_sections:
            if book_count > self.book_number:   break
            book_count += 1
            a_tags = book.find_all('a', attrs={'class':'a-link-normal'})
            book_details_link = book_prefix[self.country] + a_tags[0]['href']
            try:
                browser.get(book_details_link)
                source = browser.page_source
                soup = BeautifulSoup(source, features='lxml')
                
                title = soup.find('span', attrs={'id':'productTitle'}).get_text().strip('\n')
                author = soup.find('a', attrs={'class':'a-link-normal contributorNameID'}).get_text()
                rating = soup.find('span', attrs={'id':'acrCustomerReviewText'}).get_text().replace(' ratings', '')
                stars = soup.find('span', attrs={'class': 'reviewCountTextLinkedHistogram noUnderline'})['title'].replace(' out of 5 stars', '')
                div = soup.find('div', attrs={'cel_widget_id':'dpx-detail-bullets_csm_instrumentation_wrapper'})
                ul_tags = div.find_all('ul', attrs={'class':'a-unordered-list a-nostyle a-vertical a-spacing-none detail-bullet-list'})
                
                details = {}
                for name in self.sub_names: details[name] = self.sub_names[name]
                if 'Title'          in self.data_fields:    details['Title']        = title
                if 'Web-Link'       in self.data_fields:    details['Web-Link']     = book_details_link
                if 'Author'         in self.data_fields:    details['Author']       = author
                if 'Ratings'        in self.data_fields:    details['Ratings']      = rating
                if 'Stars'          in self.data_fields:    details['Stars']        = stars

                li_list = ul_tags[0].find_all('li')
                for li in li_list:
                    span = li.find('span').find_all('span')
                    heading = str( span[0].get_text() ).split(':')[0].strip('\n')                        
                    data = span[1].get_text()
                    try:
                        if heading == 'Publisher':
                            data = data.split('(')
                            details[heading] = data[0]
                            details['Publication date'] = data[1].replace(')', '')
                        else:
                            details[heading] = data
                    except:
                        details[heading] = data

                li = ul_tags[1].find('li')
                span = li.find('span').get_text().replace('Best Sellers Rank:', '')
                span = span.replace(' (See Top 100 in Kindle Store)', '')
                span = span.replace('\n\n\n\n','')
                span = span.replace('\n\n','\n')
                details['Best Sellers Rank'] = span
                # print('K-', self.count, '. ', details, '\n')
                # print(self.count, end=" ")
                books.append(details)
                # self.count += 1
            except:
                continue
        # Saving in Excel File  
        self.write_to_excel(self.kindle_filename, books)

    def create_excel_file(self, category_name, filename):
        # creating new excel file
        workbook = xlsxwriter.Workbook(filename)
        workbook.add_worksheet(category_name)
        workbook.close()
        workbook = op.load_workbook(filename, False)
        worksheet = workbook[category_name]
        worksheet.append(self.headers())
        workbook.save(filename)
        workbook.close()

    def write_to_excel(self, filename, books=[]):
        workbook = op.load_workbook(filename, False)
        worksheet = workbook[self.sub_names['category']]
        for book in books:
            data = []
            for data_field in self.data_fields:
                try:
                    if data_field == 'Best Sellers Rank':
                        values = book[data_field].split('\n')
                        for v in values: data.append(v)
                    else:   data.append(book[data_field])
                except: data.append('N/A')
            worksheet.append(data)
        workbook.save(filename)
        workbook.close()

    def headers(self):
        header = []
        for data_field in self.data_fields:
            header.append(data_field)
        return header


def selected_countries():
    countries =  pd.read_excel('../category_list/countries.xlsx', 'countries')
    datas = countries['countries']
    country_list = []
    for i in datas.index:
            if datas[i] is not nan:    
                country_list.append(datas[i])
    return country_list

def main():
    countries = selected_countries()
    print(countries)
    for country in countries:
        try:
            kind = kindle(country)
            kind.scrape_category()
            done = False
            while not done:
                time.sleep(5)
                # print('Work Done - ', kind.work_done)
                for val in kind.work_done:
                    if not val:
                        done = False
                        break
                    else:   done = True
        except: print(country, '-List Not Found')

if __name__ == '__main__':
   main()
